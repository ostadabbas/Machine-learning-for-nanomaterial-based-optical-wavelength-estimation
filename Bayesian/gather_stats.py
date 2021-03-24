## Bayesian inference, part 1:

import openpyxl
import numpy as np
import pandas
import json

trans = "Transmittance.xlsx"

min_count = 10

def load(fname):
    wb = openpyxl.load_workbook(fname)
    return wb

def extract_vals(ws):
    df = pandas.DataFrame(ws.values)
    dfv = df.loc[1:,1:]
    
    # get the wavelengths
    cnt = dfv.count(axis=1) >= min_count
    wavelengths = df[0][1:]
    wavelengths = wavelengths[cnt]
    
    
    # for each wavelength, get mean/std
    dfv = dfv[cnt]
    means = dfv.mean(axis=1,skipna=True)
    stds = dfv.std(axis=1,skipna=True)

    return np.array(wavelengths),np.array(means),np.array(stds)

def combine(wb, sensors):
    # extract all values
    allvals = [extract_vals(wb[s]) for s in sensors]
    
    # find the set of wavelengths common to all sensors
    wavelengths = set(allvals[0][0])
    for wls,mns,stds in allvals:
        wavelengths = wavelengths & set(wls)
    wavelengths = list(wavelengths)
    
    sz_w = len(wavelengths)
    sz_s = len(sensors)
    all_means = np.zeros((sz_w,sz_s))
    all_stds = np.zeros((sz_w,sz_s))
    for idx,vals in enumerate(allvals):
        wls,mns,stds = vals
        wlsD = {w:idx for idx,w in enumerate(wls)}
        usew=[wlsD[w] for w in wavelengths]          
        
        all_means[:,idx] = mns[usew]
        all_stds[:,idx] = stds[usew]
    return wavelengths,all_means,all_stds
def load_all(fname):
    wb = load(fname)
    sensors = wb.sheetnames 
    waves,mns,stds = combine(wb,sensors)
    return sensors,waves,mns,stds

def save_model(fname,jsonfname):
    sensors,waves,mns,stds = load_all(fname)
    outd = { 'sensors': sensors,
            'wavelengths': waves,
            'means': mns.tolist(),
            'stds': stds.tolist(),
            }
    with open(jsonfname, 'w') as f:
        json.dump(outd,f, sort_keys=True, indent=4)

def load_test(fname):
    wb = load(fname)
    tfreqs = wb.sheetnames
    def proc():
        for f in tfreqs:
            df = pandas.DataFrame(wb[f].values)
            data = np.array(df.dropna())
            # data = np.insert(data, 0, float(f), axis=1)
            yield data
    vals = np.array(list(proc()))
    vals = vals.reshape((-1,vals.shape[2]))
    
    return vals
    

if __name__ == "__main__":
    save_model("Transmittance.xlsx", "trans.json")
    
    # vals = load_test('TestT.xlsx')
    # np.savetxt('testT.csv',vals, delimiter=',')